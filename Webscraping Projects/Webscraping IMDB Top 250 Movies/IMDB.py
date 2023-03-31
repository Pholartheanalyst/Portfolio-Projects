from bs4 import BeautifulSoup
import requests
import openpyxl
import pandas as pd
import numpy as np



#create a new excel book to store the scraped data
excel = openpyxl.Workbook()
#check how many sheets the excel workbook has
'''print(IMDB_data.sheetnames)'''
#getting the active sheet
sheet = excel.active
#renaming the active sheet
sheet.title ='Scraped_IMDB_Data'
print(excel.sheetnames)

#writing the columns names into the excel sheet 
sheet.append(['Title', 'Year', 'Rating', 'Rank'])


try:
    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()

    soup=BeautifulSoup(source.text, 'html.parser') 

    movies = soup.find('tbody', class_='lister-list').find_all('tr')

    for m in movies:
        Title = m.find('td', class_='titleColumn').a.text
        Year = m.find('td', class_='titleColumn').find('span', class_='secondaryInfo').text.strip('()')
        Rating = m.find('td', class_='ratingColumn imdbRating').strong.text
        Rank = m.find('td', class_='titleColumn').get_text().split('.')[0]

        Details = [Title, Year, Rating, Rank]

        sheet.append(Details)


except Exception as e:
    print(e)

excel.save('IMDB_WebScraped_Data.csv')

df = pd.DataFrame(columns=['Title', 'Year', 'Rating','Rank'])
df.head()